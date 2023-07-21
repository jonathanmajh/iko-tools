import csv
# from ctypes import alignment
# from textwrap import fill
import requests
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, GradientFill
from openpyxl.formatting.rule import IconSetRule

# APIKEYS should be set as a user local environment variable for security reasons
# MAS is available over the internet
APIKEY = os.getenv('MASPROD')

# --------------------------------
# variables to change
month = 6
year = 2023
# variables to change
# --------------------------------
current_month = year * 12 + month

# must be in the same order as the KPIs since order is involved in overall average calculations
ENDPOINTS = {
    'IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME': 'VH & H Critical Asset PM Completed On Time Trend',
    'IKO_KPI_MAXIMOVSKRONOSHOURS': 'Maximo Lbr Hrs vs Kronos Lb Hrs',
    'IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST': 'Issued Items Listed in Spare Part List',
    'IKO_KPIWOW1STWHY': "Work Orders with 5 Why's",
    'IKO_KPI_MROLINEFROMMAXIMO': 'MRO PR Lines from Maximo',
    'IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY': 'Asset Spare Parts Count',
    'IKO_KPI_CYCLECOUNT': 'Cycle Count',
    'IKO_KPI_SCHEDULEDWORKORDERS': 'Scheduled Work Orders',
    'IKO_KPI_WONOTCREATEDBYPLANNER': 'WO Not Created by Planner',
    'IKO_KPI_JOBPLANCREATED': 'Job Plans Created/Updated',
    'IKO_KPI_LABORHOURCHARGEDTOLLCA': 'Labor Hrs Charged to Lowest Child Assets',
}

averages = {
    'IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME': [0, 0, 0.0],
    'IKO_KPI_MAXIMOVSKRONOSHOURS': [0, 0, 0.0],
    'IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST': [0, 0, 0.0],
    'IKO_KPIWOW1STWHY': [0, 0, 0.0],
    'IKO_KPI_MROLINEFROMMAXIMO': [0, 0, 0.0],
    'IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY': [0, 0, 0.0],
    'IKO_KPI_CYCLECOUNT': [0, 0, 0.0],
    'IKO_KPI_SCHEDULEDWORKORDERS': [0, 0, 0.0],
    'IKO_KPI_WONOTCREATEDBYPLANNER': [0, 0, 0.0],
    'IKO_KPI_JOBPLANCREATED': [0, 0, 0.0],
    'IKO_KPI_LABORHOURCHARGEDTOLLCA': [0, 1, 0.0],
}

overall = {'up': 0, 'side': 0, 'down': 0}

ENDPOINT_URL = list(ENDPOINTS)

PMONTIME = 'IKO_KPI_PMCOMPLETEDONTIME'

SITES = {
    'GI': 'Madoc',
    'GE': 'Ashcroft',
    'GS': 'Sylacauga',
    'BA': 'Calgary',
    'GV': 'Hillsboro',
    'GH': 'Hawkesbury',
    'AA': 'IKO Brampton',
    'GJ': 'CRC Toronto',
    'CA': 'Kankakee',
    'GC': 'Sumas',
    'GK': 'IG Brampton',
    'CAM': 'Appley Bridge',
    'BL': 'Hagerstown',
    'RAM': 'Alconbury',
    'GP': 'CRC Brampton',
    'GM': 'IG High River',
    'COM': 'Combronde',
    'GR': 'Bramcal',
    'GX': 'MaxiMix',
    'ANT': 'Antwerp'
}

MONTHS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]

URL = [
    'https://prod.manage.prod.iko.max-it-eam.com/maximo/api/script/',
    '?site=',
    '',
]

results = [
    ["Site"],
    ["KPI metric"],
    [],  # spacer
    ['Average'],
    [],  # spacer
    [MONTHS[month - 1]],
    [MONTHS[month - 2]],
    [MONTHS[month - 3]],
    [MONTHS[month - 4]],
    [MONTHS[month - 5]],
    [MONTHS[month - 6]],
    [MONTHS[month - 7]],
    [MONTHS[month - 8]],
    [MONTHS[month - 9]],
    [MONTHS[month - 10]],
    [MONTHS[month - 11]],
    [MONTHS[month % 12]],
]


def request_wrapper(url):
    header = {'apikey': APIKEY}
    r = requests.get(url, headers = header)
    if (r.status_code != 200):
        print(f'Error: {r.status_code}')
        print(f'Error: {url}')
        result = {'target': -1}
        return result
    else:
        r = r.json()
        if (not 'info' in r):
            r['info'] = []
        return r


def level(table):
    columns = len(table[0])
    for j in range(len(table)):
        while (len(table[j]) < columns):
            table[j].append('-')
    return table


def zero(table):
    columns = len(table[0])
    for j in range(len(table)):
        while (len(table[j]) < columns):
            table[j].append(0)
    return table


def compare_values(target, results, ws, ws_write_col):
    # shouldnt need to return ws back
    # ws.cell(row=3, column=ws_write_col, value=target)
    if (results[3][-1] != '-' and target != '-'):
        if ((target <= results[3][-1])):
            ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = green
        elif ((target - results[3][-1]) < (0.2 * target)):
            ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = yellow
        else:
            ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = red
        ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = '0%'
    if (results[5][-1] != '-'):
        if ((target <= results[5][-1])):
            ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = green
        elif ((target - results[5][-1]) < (0.2 * target)):
            ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = yellow
        else:
            ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = red
        ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = '0%'
        # record value for overall average
        averages[ENDPOINT_URL[ws_write_col - 8]][0] += results[5][-1]
        averages[ENDPOINT_URL[ws_write_col - 8]][1] += 1
    if ((results[5][-1] == '-') or (results[3][-1] == '-')):  # blank is meh
        ws.cell(row=3 * site_i + 3, column=ws_write_col, value=0)
        overall['side'] += 1
    elif (abs(results[3][-1] - results[5][-1]) <
          0.005):  # very small change is meh
        ws.cell(row=3 * site_i + 3, column=ws_write_col, value=0)
        overall['side'] += 1
    elif (results[3][-1] < results[5][-1]):  # above average
        ws.cell(row=3 * site_i + 3, column=ws_write_col, value=1)
        overall['up'] += 1
    else:  # below average
        ws.cell(row=3 * site_i + 3, column=ws_write_col, value=-1)
        overall['down'] += 1
    ws.cell(row=3 * site_i + 1, column=ws_write_col).alignment = center
    ws.cell(row=3 * site_i + 2, column=ws_write_col).alignment = center
    ws.cell(row=3 * site_i + 3, column=ws_write_col).alignment = center


def generic_result_reader(url, results, ws, ws_write_col):
    result = request_wrapper(url)
    if (result == -1):
        for j in range(len(results) - 2):
            results[j + 2].append("-")
    else:
        overall_percent = 0.0
        for item in result['info']:
            monthdate = item['year'] * 12 + item['month']
            results[5 + (current_month - monthdate)].append(item['percentage'])
            overall_percent += item['percentage']
        if (len(result['info']) > 0):
            results[3].append(overall_percent / len(result['info']))
    results = level(results)
    ws.cell(row=3 * site_i + 1, column=ws_write_col, value=results[3][-1])
    ws.cell(row=3 * site_i + 2, column=ws_write_col, value=results[5][-1])
    compare_values(result['target'], results, ws, ws_write_col)


wb = Workbook()
ws = wb.active
site_i = 0

#text alignment
center = Alignment(horizontal='center', vertical='center')
ninety_center = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=90)

#cell colors
green = PatternFill(start_color='00C6EFCE',
                    end_color='00C6EFCE',
                    fill_type='solid')
yellow = PatternFill(start_color='00FFEB9C',
                     end_color='00FFEB9C',
                     fill_type='solid')
red = PatternFill(start_color='00FFC7CE',
                  end_color='00FFC7CE',
                  fill_type='solid')
white = PatternFill(fill_type=None,
                    start_color='FFFFFFFF',
                    end_color='FF000000')
#border styles
thin = Side(border_style="thin", color="000000")
double = Side(border_style="medium", color="000000")

# format workbook
ws.append([])
# should be generated from above list
ws.append([
    '',
    '',
    '',
    '',
    '',
    'KPI Summary',
    '',
    'VH & H Critical Asset PM Completed On Time',
    'Maximo Lbr Hrs vs Timesheet Hrs',
    'Issued Items Listed in Spare Part List',
    "Work Orders with 5 Why's Completed",
    'MRO PR Lines from Maximo',
    'Asset Spare Parts Count with Criticality',
    'Cycle Count',
    'Scheduled Work Orders',
    'WO Not Created by Planner',
    'Job Plans Created/Updated',
    # 'Labor Hrs Charged to Lowest Child Assets',
])

ws.append([
    '',
    '',
    '',
    '',
    '',
    'Target',
    '',
    '95%',
    '80%',
    '80%',
    '12',
    '80%',
    '-',
    '10%',
    '80%',
    '90%',
    25,
])

ws.cell(row=2, column=8).fill = PatternFill(start_color='FF6565',
                                            end_color='FF6565',
                                            fill_type='solid')
ws.cell(row=2, column=9).fill = PatternFill(start_color='FF6565',
                                            end_color='FF6565',
                                            fill_type='solid')
ws.cell(row=2, column=10).fill = PatternFill(start_color='FABF8F',
                                             end_color='FABF8F',
                                             fill_type='solid')
ws['K2'].fill = GradientFill(stop=("FABF8F", "B1A0C7"))
ws.cell(row=2, column=12).fill = PatternFill(start_color='B1A0C7',
                                             end_color='B1A0C7',
                                             fill_type='solid')
ws.cell(row=2, column=13).fill = PatternFill(start_color='FABF8F',
                                             end_color='FABF8F',
                                             fill_type='solid')
ws.cell(row=2, column=14).fill = PatternFill(start_color='B1A0C7',
                                             end_color='B1A0C7',
                                             fill_type='solid')
ws.cell(row=2, column=15).fill = PatternFill(start_color='00B0F0',
                                             end_color='00B0F0',
                                             fill_type='solid')
ws.cell(row=2, column=16).fill = PatternFill(start_color='00B0F0',
                                             end_color='00B0F0',
                                             fill_type='solid')
ws['Q2'].fill = GradientFill(stop=("FABF8F", "00B0F0"))

ws.cell(row=4, column=21, value='Planner / Scheduler')
ws.cell(row=4, column=20).fill = PatternFill(start_color='00B0F0',
                                             end_color='00B0F0',
                                             fill_type='solid')
ws.cell(row=5, column=21, value='Inventory Coordinator')
ws.cell(row=5, column=20).fill = PatternFill(start_color='B1A0C7',
                                             end_color='B1A0C7',
                                             fill_type='solid')
ws.cell(row=6, column=21, value='Reliability Engineer')
ws.cell(row=6, column=20).fill = PatternFill(start_color='FABF8F',
                                             end_color='FABF8F',
                                             fill_type='solid')
ws.cell(row=7, column=21, value='Maintenance Manager')
ws.cell(row=7, column=20).fill = PatternFill(start_color='FF6565',
                                             end_color='FF6565',
                                             fill_type='solid')

ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)

for cell in range(13):
    ws.cell(row=2, column=6 + cell).alignment = Alignment(text_rotation=45)
    
ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)

for cell in range(12):
    ws.cell(row=3, column=6 + cell).border = Border(bottom=double, top=double)
    ws.cell(row=2, column=6 + cell).border = Border(top=double)
    ws.cell(row=3, column=6 + cell).alignment = center

ws.cell(row=4, column=2, value='GRANULES')
ws.cell(row=4, column=2).alignment = ninety_center
ws.cell(row=4, column=2).border = Border(top=double,
                                         left=double,
                                         right=double,
                                         bottom=double)
ws.cell(row=4, column=2).font = Font(name='Calibri', size=12, bold=True)
ws.merge_cells(start_row=4, start_column=2, end_row=9, end_column=2)
ws.column_dimensions['B'].width = 2
ws.cell(row=4, column=3).border = Border(top=double)
ws.cell(row=4, column=4).border = Border(top=double)
ws.cell(row=4, column=5).border = Border(top=double)

ws.cell(row=10, column=3, value='SHINGLE')
ws.cell(row=10, column=3).alignment = ninety_center
ws.cell(row=10, column=3).border = Border(top=double,
                                          left=double,
                                          right=double,
                                          bottom=double)
ws.cell(row=10, column=3).font = Font(name='Calibri', size=12, bold=True)
ws.merge_cells(start_row=10, start_column=3, end_row=33, end_column=3)
ws.column_dimensions['C'].width = 2
ws.cell(row=10, column=4).border = Border(top=double)
ws.cell(row=10, column=5).border = Border(top=double)

ws.cell(row=28, column=4, value='MEMBRANES')
ws.cell(row=28, column=4).alignment = ninety_center
ws.cell(row=28, column=4).border = Border(top=double,
                                          left=double,
                                          right=double,
                                          bottom=double)
ws.cell(row=28, column=4).font = Font(name='Calibri', size=12, bold=True)
ws.merge_cells(start_row=28, start_column=4, end_row=42, end_column=4)
ws.column_dimensions['D'].width = 2
ws.cell(row=28, column=5).border = Border(top=double)

ws.cell(row=40, column=5, value='ISO')
ws.cell(row=40, column=5).alignment = ninety_center
ws.cell(row=40, column=5).border = Border(top=double,
                                          left=double,
                                          right=double,
                                          bottom=double)
ws.cell(row=40, column=5).font = Font(name='Calibri', size=12, bold=True)
ws.merge_cells(start_row=40, start_column=5, end_row=54, end_column=5)
ws.column_dimensions['E'].width = 2
ws.cell(row=2, column=6).border = Border(left=double,
                                         bottom=double,
                                         top=double)
ws.cell(row=3, column=6).border = Border(left=double,
                                         bottom=double,
                                         top=double)
ws.cell(row=2, column=17).border = Border(right=double,
                                          bottom=double,
                                          top=double)
ws.cell(row=3, column=17).border = Border(right=double,
                                          bottom=double,
                                          top=double)
ws.column_dimensions['F'].width = 13
# iconset test
rule = IconSetRule('3Arrows',
                   'num', [-1, 0, 1],
                   showValue=False,
                   percent=None,
                   reverse=None)

for site in SITES:
    ws_write_col = 8
    site_i += 1
    ws.cell(row=3 * site_i + 2, column=6, value=SITES[site])
    ws.cell(row=3 * site_i + 1, column=6).alignment = center
    ws.cell(row=3 * site_i + 1, column=6).border = Border(left=double,
                                                          top=double)
    ws.cell(row=3 * site_i + 2, column=6).border = Border(left=double)
    ws.cell(row=3 * site_i + 3, column=6).border = Border(left=double)
    ws.cell(row=3 * site_i + 1, column=7, value='Avg')
    ws.cell(row=3 * site_i + 1, column=7).alignment = center
    ws.cell(row=3 * site_i + 2, column=7, value=MONTHS[month - 1])
    ws.cell(row=3 * site_i + 2, column=7).alignment = center
    ws.cell(row=3 * site_i + 3, column=7, value='Trend')
    ws.cell(row=3 * site_i + 3, column=7).alignment = center
    for cell in range(11):
        ws.cell(row=3 * site_i + 3,
                column=7 + cell).border = Border(bottom=double)
    ws.cell(row=3 * site_i + 1, column=17).border = Border(right=double,
                                                           top=double)
    ws.cell(row=3 * site_i + 2, column=17).border = Border(right=double)
    ws.cell(row=3 * site_i + 3, column=17).border = Border(right=double)
    ws.conditional_formatting.add(f'H{3*site_i +3}:T{3*site_i +3}', rule)

    # pm on time
    results[0].append(site)
    results[1].append('IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME')
    print(f'IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME : {site}')
    result = request_wrapper(
        f'{URL[0]}IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME{URL[1]}{site}{URL[2]}'
    )
    if (result == -1):
        for j in range(len(results) - 2):
            results[j + 2].append("-")
    else:
        allcount = {}
        ontimecount = {}
        percentage = {}
        overall_percent = 0.0
        for item in result['info']:
            monthdate = item['year'] * 12 + item['month']
            if (monthdate in allcount):
                allcount[monthdate] += item['allcount']
                ontimecount[monthdate] += item['ontimecount']
            else:
                allcount[monthdate] = item['allcount']
                ontimecount[monthdate] = item['ontimecount']
        for key in allcount.keys():
            results[5 + (current_month - key)].append(ontimecount[key] /
                                                      allcount[key])
            overall_percent += ontimecount[key] / allcount[key]
        if (len(allcount.keys()) > 0):
            results[3].append(overall_percent / len(allcount.keys()))
        else:
            print('Falling back to All PM Completed on Time KPI')
            result = request_wrapper(
                f'{URL[0]}{PMONTIME}{URL[1]}{site}{URL[2]}')
            if (result != -1):
                for item in result['info']:
                    monthdate = item['year'] * 12 + item['month']
                    results[5 + (current_month - monthdate)].append(
                        item['percentage'])
                    overall_percent += item['percentage']
                if (len(result['info']) > 0 and overall_percent > 0.001):
                    results[3].append(overall_percent / len(result['info']))
    results = level(results)
    ws.cell(row=3 * site_i + 1, column=ws_write_col, value=results[3][-1])
    ws.cell(row=3 * site_i + 2, column=ws_write_col, value=results[5][-1])
    compare_values(result['target'], results, ws, ws_write_col)

    ws_write_col += 1
    # kronos vs maximo
    results[0].append(site)
    results[1].append('IKO_KPI_MAXIMOVSKRONOSHOURS')
    print(f'IKO_KPI_MAXIMOVSKRONOSHOURS : {site}')
    generic_result_reader(
        f'{URL[0]}IKO_KPI_MAXIMOVSKRONOSHOURS{URL[1]}{site}{URL[2]}', results,
        ws, ws_write_col)

    ws_write_col += 1
    # IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST
    results[0].append(site)
    results[1].append('IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST')
    print(f'IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST : {site}')
    generic_result_reader(
        f'{URL[0]}IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST{URL[1]}{site}{URL[2]}',
        results, ws, ws_write_col)

    ws_write_col += 1
    # WO With Why
    results[0].append(site)
    results[1].append('IKO_KPIWOW1STWHY')
    print(f'IKO_KPIWOW1STWHY : {site}')
    result = request_wrapper(f'{URL[0]}IKO_KPIWOW1STWHY{URL[1]}{site}{URL[2]}')
    if (result == -1):
        for j in range(len(results) - 2):
            results[j + 2].append("-")
    else:
        allcount = {}
        overall_percent = 0.0
        for item in result['info']:
            monthdate = item['year'] * 12 + item['month']
            overall_percent += item['whycompletecount']
            if (monthdate in allcount):
                allcount[monthdate] += item['whycompletecount']
            else:
                allcount[monthdate] = item['whycompletecount']
        for monthdate in allcount.keys():
            results[5 + (current_month - monthdate)].append(
                allcount[monthdate])
        if (len(result['info']) > 0):
            results[3].append(overall_percent / len(allcount.keys()))
    results = level(results)
    ws.cell(row=3 * site_i + 1, column=ws_write_col, value=results[3][-1])
    ws.cell(row=3 * site_i + 2, column=ws_write_col, value=results[5][-1])
    compare_values(result['target'], results, ws, ws_write_col)
    ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = '#,##0'
    ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = '#,##0'
    ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = white
    ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = white
    # generic_result_reader(f'{URL[0]}{ENDPOINT_URL[2]}{URL[1]}{site}{URL[2]}',
    #                       results, ws, ws_write_col)

    ws_write_col += 1
    # IKO_KPI_MROLINEFROMMAXIMO
    results[0].append(site)
    results[1].append('IKO_KPI_MROLINEFROMMAXIMO')
    print(f'IKO_KPI_MROLINEFROMMAXIMO : {site}')
    generic_result_reader(
        f'{URL[0]}IKO_KPI_MROLINEFROMMAXIMO{URL[1]}{site}{URL[2]}', results,
        ws, ws_write_col)

    ws_write_col += 1
    # IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY
    results[0].append(site)
    results[1].append('IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY')
    print(f'IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY : {site}')
    result = request_wrapper(
        f'{URL[0]}IKO_API_ASSETSPAREPARTCOUNTWITHCRITICALITY{URL[1]}{site}{URL[2]}'
    )
    if (result == -1):
        for j in range(len(results) - 2):
            results[j + 2].append("-")
    else:
        allcount = {}
        overall_percent = 0.0
        for item in result['info']:
            monthdate = item['year'] * 12 + item['month']
            overall_percent += item['sparepartcount']
            if (monthdate in allcount):
                allcount[monthdate] += item['sparepartcount']
            else:
                allcount[monthdate] = item['sparepartcount']
        for monthdate in allcount.keys():
            results[5 + (current_month - monthdate)].append(
                allcount[monthdate])
        if (len(result['info']) > 0):
            results[3].append(overall_percent / len(allcount.keys()))
    results = level(results)
    ws.cell(row=3 * site_i + 1, column=ws_write_col, value=results[3][-1])
    ws.cell(row=3 * site_i + 2, column=ws_write_col, value=results[5][-1])
    compare_values(result['target'], results, ws, ws_write_col)
    ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = '#,##0'
    ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = '#,##0'
    ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = white
    ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = white

    ws_write_col += 1
    # IKO_KPI_CYCLECOUNT
    results[0].append(site)
    results[1].append('IKO_KPI_CYCLECOUNT')
    print(f'IKO_KPI_CYCLECOUNT : {site}')
    generic_result_reader(f'{URL[0]}IKO_KPI_CYCLECOUNT{URL[1]}{site}{URL[2]}',
                          results, ws, ws_write_col)

    ws_write_col += 1
    # IKO_KPI_SCHEDULEDWORKORDERS
    results[0].append(site)
    results[1].append('IKO_KPI_SCHEDULEDWORKORDERS')
    print(f'IKO_KPI_SCHEDULEDWORKORDERS : {site}')
    generic_result_reader(
        f'{URL[0]}IKO_KPI_SCHEDULEDWORKORDERS{URL[1]}{site}{URL[2]}', results,
        ws, ws_write_col)

    ws_write_col += 1
    # IKO_KPI_WONOTCREATEDBYPLANNER
    results[0].append(site)
    results[1].append('IKO_KPI_WONOTCREATEDBYPLANNER')
    print(f'IKO_KPI_WONOTCREATEDBYPLANNER : {site}')
    generic_result_reader(
        f'{URL[0]}IKO_KPI_WONOTCREATEDBYPLANNER{URL[1]}{site}{URL[2]}',
        results, ws, ws_write_col)

    ws_write_col += 1
    # IKO_KPI_JOBPLANCREATED
    results[0].append(site)
    results[1].append('IKO_KPI_JOBPLANCREATED')
    print(f'IKO_KPI_JOBPLANCREATED : {site}')
    result = request_wrapper(
        f'{URL[0]}IKO_KPI_JOBPLANCREATED{URL[1]}{site}{URL[2]}')
    if (result == -1):
        for j in range(len(results) - 2):
            results[j + 2].append("-")
    else:
        overall_percent = 0.0
        for item in result['info']:
            monthdate = item['year'] * 12 + item['month']
            results[5 + (current_month - monthdate)].append(item['total'])
            overall_percent += item['total']
        if (len(result['info']) > 0):
            results[3].append(overall_percent / len(result['info']))
    results = level(results)
    ws.cell(row=3 * site_i + 1, column=ws_write_col, value=results[3][-1])
    ws.cell(row=3 * site_i + 2, column=ws_write_col, value=results[5][-1])
    compare_values(result['target'], results, ws, ws_write_col)
    ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = '#,##0'
    ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = '#,##0'

    # ws_write_col += 1
    # # IKO_KPI_LABORHOURCHARGEDTOLLCA
    # results[0].append(site)
    # results[1].append(ENDPOINT_URL[3])
    # print(f'{ENDPOINT_URL[3]} : {site}')
    # generic_result_reader(f'{URL[0]}{ENDPOINT_URL[3]}{URL[1]}{site}{URL[2]}',
    #                       results, ws, ws_write_col)

for cell in range(12):
    ws.cell(row=64, column=6 + cell).border = Border(top=double)

i = 0
for average in averages:
    ws.cell(row=70,
            column=8 + i,
            value=averages[average][0] / averages[average][1])
    i += 1

ws.cell(row=72, column=6, value='Up')
ws.cell(row=72, column=7, value=overall['up'])
ws.cell(row=73, column=6, value='Same')
ws.cell(row=73, column=7, value=overall['side'])
ws.cell(row=74, column=6, value='Down')
ws.cell(row=74, column=7, value=overall['down'])

# write
with open(
        f"C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\KPI\\{date.today().isoformat()}_KPI.csv",
        "w",
        newline="",
) as f:
    writer = csv.writer(f)
    writer.writerows(results)

wb.save(
    filename=
    f"C:\\Users\\majona\\Documents\\Code\\\iko-tools\\Python\\KPI\\{date.today().isoformat()}_KPI.xlsx"
)
