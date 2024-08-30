import csv

# from ctypes import alignment
# from textwrap import fill
import requests
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment, GradientFill
from openpyxl.formatting.rule import IconSetRule

# APIKEYS should be set as a user local environment variable for security reasons
# MAS is available over the internet
APIKEY = os.getenv("MASPROD")

# --------------------------------
# variables to change
month = 7
year = 2024
# variables to change
# --------------------------------

# must be in the same order as the KPIs since order is involved in overall average calculations
# 0 = percentage, 1 = add everything
ENDPOINTS = {
    "IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME": [
        "VH & H Critical Asset PM Completed On Time Trend",
        0,
    ],
    "IKO_KPI_MAXIMOVSKRONOSHOURS": ["Maximo Lbr Hrs vs Timesheet Hrs", 0],
    "IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST": [
        "Spare Parts Issued to LLCA on Spare Part List",
        0,
    ],
    "IKO_KPI_CYCLECOUNT": ["Cycle Count & Discrepancies", 0],
    "IKO_KPI_SPAREISSUEWITHOUTLLCA": ["Spare Parts Issued Without LLCA Work Order", 1],
    "IKO_KPI_ASSETSPAREPARTCOUNTWITHCRITICALITY": [
        "Asset Spare Part Count w/ Criticality",
        1,
    ],
    "IKO_KPI_SPAREPURCHASED": ["Spare Parts Purchased on Spare Part List", 0],
    "IKO_KPI_MROLINEFROMMAXIMO": ["MRO PR Lines from Maximo", 0],
    "IKO_KPI_SDT": ["SDT PM Implementation", 0],
    "IKO_KPIWOW1STWHY": ["Work Orders with 5 Why's Completed", 1],
    # "IKO_KPI_SCHEDULEDWORKORDERS": ["Scheduled Work Orders", 0],
    # "IKO_KPI_WONOTCREATEDBYPLANNER": ["WO Not Created by Planner", 0],
    # "IKO_KPI_JOBPLANCREATED": ["Job Plans Created/Updated", 0],
}

order = [
    "IKO_KPI_HVHCRITICALASSETPMCOMPLETEDONTIME",
    "IKO_KPI_MAXIMOVSKRONOSHOURS",
    "IKO_KPI_ISSUEDITEMLISTEDINSPAREPARTLIST", #ok
    "IKO_KPI_SPAREPURCHASED", #ok
    "IKO_KPI_SPAREISSUEWITHOUTLLCA", #ok
    "IKO_KPI_CYCLECOUNT", #ok
    "IKO_KPI_MROLINEFROMMAXIMO", #ok
    "IKO_KPI_ASSETSPAREPARTCOUNTWITHCRITICALITY", #ok
    "IKO_KPIWOW1STWHY", #ok
    "IKO_KPI_SDT", #ok
    # "IKO_KPI_SCHEDULEDWORKORDERS",
    # "IKO_KPI_WONOTCREATEDBYPLANNER",
    # "IKO_KPI_JOBPLANCREATED",
]

# used to track kpi value for the month and averages
# averages[kpi point][siteid] = [0.0, 0.0, 0.0, 0.0]
# number of months, total / average of all months, current month, target
averages = {}

# used for calculating average for org
# [total, denominator]
overall = {}

ENDPOINT_URL = list(ENDPOINTS)

SITES = {
    # granule
    "GE": "Ashcroft",
    "GI": "Madoc",
    # shingle
    "GS": "Sylacauga",
    "BA": "Calgary",
    "GV": "Hillsboro",
    "GH": "Hawkesbury",
    "AA": "IKO Brampton",
    # mod
    "CA": "Kankakee",
    "GC": "Sumas",
    "GK": "IG Brampton",
    "CAM": "Appley Bridge",
    "ANT": "Antwerp",
    #  ISO
    "BL": "Hagerstown",
    "RAM": "Alconbury",
    "GP": "CRC Brampton",
    "GM": "IG High River",
    "COM": "Combronde",
    "KLU": "Klundert",
    # Other
    "GR": "Bramcal",
    "GX": "MaxiMix",
}

MONTHS = [
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
]

URL = "https://prod.manage.prod.iko.max-it-eam.com/maximo/api/script/"

def request_wrapper(url):
    header = {"apikey": APIKEY}
    r = requests.get(url, headers=header)
    if r.status_code != 200:
        print(f"Error: {r.status_code}")
        print(f"Error: {url}")
        result = {"target": -1}
        return result
    else:
        r = r.json()
        return r


def kpi_value(kpi_type: int, values):
    if kpi_type == 0:
        return values["percentage"]
    else:
        temp = 0
        for key, value in values.items():
            if key not in ['year', 'month', 'siteid', 'siteDesc', 'target', 'priority']:
                temp += value
        return temp


def compare_values(target, value):
    # shouldnt need to return ws back
    # ws.cell(row=3, column=ws_write_col, value=target)
    if target == '-':
        return white
    # for 0 target anything > 0 is bad
    if target == 0:
        if target < value:
            return red
        return green
    if target <= value:
        return green
    if (target - value) < (0.2 * target):
        return yellow
    return red

data_wb = Workbook()

def dumpData(dataWb, data, kpiName):
    headers = []
    row = []
    ws = dataWb.create_sheet(kpiName[7:35], 0)
    for column in data[0]:
        headers.append(column)
    ws.append(headers)
    for item in data:
        row = []
        for header in headers:
            row.append(item[header])
        ws.append(row)

# load data from APIs
for points in ENDPOINT_URL:
    print(points)
    result = request_wrapper(f"{URL}{points}")
    dumpData(data_wb, result, points)
    averages[points] = {}
    overall[points] = [0,0]
    for item in result:
        if item["siteid"] in ['GJ']:
            continue
        if item["siteid"] not in averages[points]:
            averages[points][item["siteid"]] = [0.0, 0.0, 0.0, 0.0]
        averages[points][item["siteid"]][0] += 1
        averages[points][item["siteid"]][1] += kpi_value(ENDPOINTS[points][1], item)
        averages[points][item["siteid"]][3] = item['target']
        if item["year"] == year and item["month"] == month:
            averages[points][item["siteid"]][2] = kpi_value(ENDPOINTS[points][1], item)
            overall[points][0] += kpi_value(ENDPOINTS[points][1], item)
    for sites in averages[points]:
        averages[points][sites][1] = averages[points][sites][1] / averages[points][sites][0]
        if (averages[points][sites][1] > 0):
            overall[points][1] += 1

wb = Workbook()
ws = wb.active
site_i = 0

# text alignment
center = Alignment(horizontal="center", vertical="center")
ninety_center = Alignment(horizontal="center", vertical="center", text_rotation=90)

# cell colors
green = PatternFill(start_color="00C6EFCE", end_color="00C6EFCE", fill_type="solid")
yellow = PatternFill(start_color="00FFEB9C", end_color="00FFEB9C", fill_type="solid")
red = PatternFill(start_color="00FFC7CE", end_color="00FFC7CE", fill_type="solid")
white = PatternFill(fill_type=None, start_color="FFFFFFFF", end_color="FF000000")
# border styles
thin = Side(border_style="thin", color="000000")
double = Side(border_style="medium", color="000000")

# format workbook
ws.append([])
# should be generated from above list
ws.append(
    [
        "",
        "",
        "",
        "",
        "",
        "KPI Summary",
        "",
        ENDPOINTS[order[0]][0],
        ENDPOINTS[order[1]][0],
        ENDPOINTS[order[2]][0],
        ENDPOINTS[order[3]][0],
        ENDPOINTS[order[4]][0],
        ENDPOINTS[order[5]][0],
        ENDPOINTS[order[6]][0],
        ENDPOINTS[order[7]][0],
        ENDPOINTS[order[8]][0],
        ENDPOINTS[order[9]][0],
    ]
)

ws.append(
    [
        "",
        "",
        "",
        "",
        "",
        "Target",
        "",
        "95%",
        "80%",
        "80%",
        "80%",
        "0",
        "10%",
        "80%",
        "-",
        "12",
        "80%",
    ]
)

# color code for responsibilities
ws.cell(row=2, column=8).fill = PatternFill(
    start_color="FF6565", end_color="FF6565", fill_type="solid"
)
ws.cell(row=2, column=9).fill = PatternFill(
    start_color="FF6565", end_color="FF6565", fill_type="solid"
)
ws["J2"].fill = GradientFill(stop=("FABF8F", "B1A0C7"))
ws["K2"].fill = GradientFill(stop=("FABF8F", "B1A0C7"))
ws["L2"].fill = GradientFill(stop=("FABF8F", "B1A0C7"))

ws.cell(row=2, column=13).fill = PatternFill(
    start_color="B1A0C7", end_color="B1A0C7", fill_type="solid"
)
ws.cell(row=2, column=14).fill = PatternFill(
    start_color="B1A0C7", end_color="B1A0C7", fill_type="solid"
)
ws.cell(row=2, column=15).fill = PatternFill(
    start_color="FABF8F", end_color="FABF8F", fill_type="solid"
)
ws.cell(row=2, column=16).fill = PatternFill(
    start_color="FABF8F", end_color="FABF8F", fill_type="solid"
)
ws.cell(row=2, column=17).fill = PatternFill(
    start_color="FABF8F", end_color="FABF8F", fill_type="solid"
)

ws.cell(row=4, column=19, value="Planner / Scheduler")
ws.cell(row=4, column=18).fill = PatternFill(
    start_color="00B0F0", end_color="00B0F0", fill_type="solid"
)
ws.cell(row=5, column=19, value="Inventory Coordinator")
ws.cell(row=5, column=18).fill = PatternFill(
    start_color="B1A0C7", end_color="B1A0C7", fill_type="solid"
)
ws.cell(row=6, column=19, value="Reliability Engineer")
ws.cell(row=6, column=18).fill = PatternFill(
    start_color="FABF8F", end_color="FABF8F", fill_type="solid"
)
ws.cell(row=7, column=19, value="Maintenance Manager")
ws.cell(row=7, column=18).fill = PatternFill(
    start_color="FF6565", end_color="FF6565", fill_type="solid"
)

ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)

for cell in range(13):
    ws.cell(row=2, column=6 + cell).alignment = Alignment(text_rotation=45)

ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=7)

for cell in range(12):
    ws.cell(row=3, column=6 + cell).border = Border(bottom=double, top=double)
    ws.cell(row=2, column=6 + cell).border = Border(top=double)
    ws.cell(row=3, column=6 + cell).alignment = center

# site grouping
ws.cell(row=4, column=2, value="GRANULES")
ws.cell(row=4, column=2).alignment = ninety_center
ws.cell(row=4, column=2).border = Border(
    top=double, left=double, right=double, bottom=double
)
ws.cell(row=4, column=2).font = Font(name="Calibri", size=12, bold=True)
ws.merge_cells(start_row=4, start_column=2, end_row=9, end_column=2)
ws.column_dimensions["B"].width = 2
ws.cell(row=4, column=3).border = Border(top=double)
ws.cell(row=4, column=4).border = Border(top=double)
ws.cell(row=4, column=5).border = Border(top=double)

ws.cell(row=10, column=3, value="SHINGLE")
ws.cell(row=10, column=3).alignment = ninety_center
ws.cell(row=10, column=3).border = Border(
    top=double, left=double, right=double, bottom=double
)
ws.cell(row=10, column=3).font = Font(name="Calibri", size=12, bold=True)
ws.merge_cells(start_row=10, start_column=3, end_row=30, end_column=3)
ws.column_dimensions["C"].width = 2
ws.cell(row=10, column=4).border = Border(top=double)
ws.cell(row=10, column=5).border = Border(top=double)

ws.cell(row=25, column=4, value="MEMBRANES")
ws.cell(row=25, column=4).alignment = ninety_center
ws.cell(row=25, column=4).border = Border(
    top=double, left=double, right=double, bottom=double
)
ws.cell(row=25, column=4).font = Font(name="Calibri", size=12, bold=True)
ws.merge_cells(start_row=25, start_column=4, end_row=39, end_column=4)
ws.column_dimensions["D"].width = 2
ws.cell(row=25, column=5).border = Border(top=double)

ws.cell(row=40, column=5, value="ISO")
ws.cell(row=40, column=5).alignment = ninety_center
ws.cell(row=40, column=5).border = Border(
    top=double, left=double, right=double, bottom=double
)
ws.cell(row=40, column=5).font = Font(name="Calibri", size=12, bold=True)
ws.merge_cells(start_row=40, start_column=5, end_row=57, end_column=5)
ws.column_dimensions["E"].width = 2
ws.cell(row=2, column=6).border = Border(left=double, bottom=double, top=double)
ws.cell(row=3, column=6).border = Border(left=double, bottom=double, top=double)
ws.cell(row=2, column=17).border = Border(right=double, bottom=double, top=double)
ws.cell(row=3, column=17).border = Border(right=double, bottom=double, top=double)
ws.column_dimensions["F"].width = 13
# iconset test
rule = IconSetRule(
    "3Arrows", "num", [-1, 0, 1], showValue=False, percent=None, reverse=None
)

for site in SITES:
    ws_write_col = 8
    site_i += 1
    ws.cell(row=3 * site_i + 2, column=6, value=SITES[site])
    ws.cell(row=3 * site_i + 1, column=6).alignment = center
    ws.cell(row=3 * site_i + 1, column=6).border = Border(left=double, top=double)
    ws.cell(row=3 * site_i + 2, column=6).border = Border(left=double)
    ws.cell(row=3 * site_i + 3, column=6).border = Border(left=double)
    ws.cell(row=3 * site_i + 1, column=7, value="Avg")
    ws.cell(row=3 * site_i + 1, column=7).alignment = center
    ws.cell(row=3 * site_i + 2, column=7, value=MONTHS[month - 1])
    ws.cell(row=3 * site_i + 2, column=7).alignment = center
    ws.cell(row=3 * site_i + 3, column=7, value="Trend")
    ws.cell(row=3 * site_i + 3, column=7).alignment = center
    for cell in range(11):
        ws.cell(row=3 * site_i + 3, column=7 + cell).border = Border(bottom=double)
    ws.cell(row=3 * site_i + 1, column=17).border = Border(right=double, top=double)
    ws.cell(row=3 * site_i + 2, column=17).border = Border(right=double)
    ws.cell(row=3 * site_i + 3, column=17).border = Border(right=double)
    ws.conditional_formatting.add(f"H{3*site_i +3}:T{3*site_i +3}", rule)

    for kpi in order:
        ws.cell(row=3 * site_i + 1, column=ws_write_col, value=averages[kpi][site][1])
        ws.cell(row=3 * site_i + 1, column=ws_write_col).fill = compare_values(averages[kpi][site][3], averages[kpi][site][1])
        ws.cell(row=3 * site_i + 2, column=ws_write_col, value=averages[kpi][site][2])
        ws.cell(row=3 * site_i + 2, column=ws_write_col).fill = compare_values(averages[kpi][site][3], averages[kpi][site][2])
        if (averages[kpi][site][1] < 2):
            ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = "0%"
            ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = "0%"
        else:
            ws.cell(row=3 * site_i + 1, column=ws_write_col).number_format = "#,##0"
            ws.cell(row=3 * site_i + 2, column=ws_write_col).number_format = "#,##0"
        if (abs(averages[kpi][site][1] - averages[kpi][site][2]) < 0.001):
            ws.cell(row=3 * site_i + 3, column=ws_write_col, value=0)
        elif (averages[kpi][site][1] < averages[kpi][site][2]):
            ws.cell(row=3 * site_i + 3, column=ws_write_col, value=1)
        else:
            ws.cell(row=3 * site_i + 3, column=ws_write_col, value=-1)
        if (averages[kpi][site][3] == 0):
            ws.cell(row=3 * site_i + 3, column=ws_write_col, value=ws.cell(row=3 * site_i + 3, column=ws_write_col).value * -1)
        ws.cell(row=3 * site_i + 1, column=ws_write_col).alignment = center
        ws.cell(row=3 * site_i + 2, column=ws_write_col).alignment = center
        ws.cell(row=3 * site_i + 3, column=ws_write_col).alignment = center
        ws_write_col += 1

for cell in range(12):
    ws.cell(row=64, column=6 + cell).border = Border(top=double)

i = 0
for kpi in order:
    ws.cell(row=70, column=8 + i, value=overall[kpi][0] / overall[kpi][1])
    i += 1

wb.save(
    filename=f"C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\KPI\\{date.today().isoformat()}_KPI.xlsx"
)

data_wb.save(
    filename=f"C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\KPI\\{date.today().isoformat()}_KPIDATA.xlsx"
)