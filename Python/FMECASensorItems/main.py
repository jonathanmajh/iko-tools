from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

filename = r'C:\Users\majona\Downloads\FMECA_Sensor_Items_inventory_spare.xlsx'
wb = load_workbook(filename, read_only=True)

ws_data = wb["SparePart"]

itemList = {}
siteList = {}
assetList = {}

for row in ws_data.iter_rows():
    if not row[0].value in itemList:
        itemList[row[0].value] = {}
    siteList = itemList[row[0].value]

    if not row[2].value in siteList:
        siteList[row[2].value] = {}
    assetList = siteList[row[2].value]

    if len(assetList) > 0 and row[1].value == None:
        continue

    if row[1].value == None:
        assetList['PO'] = row[4].value
        continue

    assetList[row[1].value] = row[4].value

ws_items = wb["CombinedSensors"]
save_path = f'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\FMECASensorItems\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
items = {}
new_wb = Workbook()
save_ws = new_wb.create_sheet('data')

for row in ws_items.iter_rows():
    if not row[0].value in items:
        items[row[0].value] = row[0].value
    else:
        print(f'{row[0].value} is duplicate skipped')
        save_ws.append([f'{row[0].value} is duplicate skipped'])
        continue
    if not row[0].value in itemList:
        print(f'{row[0].value} cannot be found')
        save_ws.append([f'{row[0].value} cannot be found'])
        continue
    for sites in itemList[row[0].value].keys():
        for assets in itemList[row[0].value][sites].keys():
            save_ws.append([row[0].value, row[1].value, None, sites, None, assets, None, itemList[row[0].value][sites][assets]])

new_wb.save(save_path)

print('done')
