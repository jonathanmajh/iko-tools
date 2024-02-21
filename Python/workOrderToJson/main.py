import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

# ------ CHANGE ME ---------
folder_path = 'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\workOrderToJson\\'
save_path = f'C:\\Users\\majona\\Documents\\Pirana\\PiranaWOConversion.xlsm'

source_wb = load_workbook(filename = save_path, read_only=True, data_only=True)
ws = source_wb["ProcessedPiranaWOs"]

upload_data = '[\n'

for row in ws.iter_rows(min_row=4):
    upload_data = upload_data + '{\n'
    upload_data = f'{upload_data}"wonum": "{row[24].value}",\n'
    upload_data = f'{upload_data}"description": "{row[26].value}",\n'
    temp = row[23].value.replace("\n", "\\n")
    temp = temp.replace('"', '\\"')
    upload_data = f'{upload_data}"description_longdescription": "{temp}",\n'
    upload_data = f'{upload_data}"siteid": "KLU",\n'
    upload_data = f'{upload_data}"orgid": "IKO-EU",\n'
    upload_data = f'{upload_data}"woclass": "WORKORDER",\n'
    upload_data = f'{upload_data}"assetnum": "{row[3].value}",\n'
    if (row[8].value != None and row[8].value != ''):
        upload_data = f'{upload_data}"supervisor": "{row[8].value}",\n'
    if (row[21].value != None and row[21].value != ''):
        upload_data = f'{upload_data}"iko_downtime": "{row[21].value}",\n'
        upload_data = f'{upload_data}"iko_desc1": "N/A",\n'
        upload_data = f'{upload_data}"jpnum": "{row[25].value}",\n'
    if (row[7].value != None and type(row[7].value) == datetime):
        upload_data = f'{upload_data}"reportdate": "{str(row[7].value).replace(" ", "T")}",\n'
    if (row[10].value != None and type(row[10].value) == datetime):
        upload_data = f'{upload_data}"schedstart": "{str(row[10].value).replace(" ", "T")}",\n'
    if (row[11].value != None and type(row[11].value) == datetime):
        upload_data = f'{upload_data}"schedfinish": "{str(row[11].value).replace(" ", "T")}",\n'
    if (row[17].value != None and type(row[17].value) == datetime):
        upload_data = f'{upload_data}"actstart": "{str(row[17].value).replace(" ", "T")}",\n'
    if (row[18].value != None and type(row[18].value) == datetime):
        upload_data = f'{upload_data}"actfinish": "{str(row[18].value).replace(" ", "T")}",\n'
    upload_data = f'{upload_data}"status": "WPLAN"\n'
    upload_data = upload_data + '},\n'

upload_data = upload_data + ']'

with open('C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\workOrderToJson\\readme.json', 'w', encoding='UTF-8') as f:
    f.write(upload_data)