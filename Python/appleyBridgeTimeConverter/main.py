import os
import glob
from datetime import datetime, timedelta, time

from openpyxl import load_workbook
from openpyxl import Workbook

from dataclasses import dataclass

# ------ CHANGE ME ---------
folder_path = "C:\\Users\\majona\\GitHub\\iko-tools\\Python\\appleyBridgeTimeConverter\\input"
save_path = f'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\appleyBridgeTimeConverter\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'
year = '2022'

@dataclass
class person:
    include: bool = False
    full_name: str = "blank"
    laborcode: str = "blank"

people = {
    "Hill A": person(True, "Hill, Alex", "UKAPAHIL"),
    "Jennings A": person(True, "Jennings, Adam", "UKAPAJEN"),
    "Ormerod C": person(True, "Ormerod, Craig", "UKAPCORM"),
    "Wignall C": person(True, "Wignall, Christopher", "UKAPCWIG"),
    "TAYLOR D": person(False, "Taylor, Dave", "UKAPDTAY"),
    "Deng E": person(True, "Deng, Eddie", "UKAPEDEN"),
    "Goodwin I": person(True, "Goodwin, Ian", "UKAPIGOO"),
    "Thompson J": person(True, "Thompson, Jack", "UKAPJTHO"),
    "Satchell L": person(True, "Satchell, Les", "UKAPLSAT"),
    "Hayward P": person(False, "Hayward, Paul", "UKAPPHAY"),
    "Pierce R": person(True, "Pierce, RaTrue", "UKAPRPIE"),
    "Coyle S": person(True, "Coyle, Steven", "UKAPSCOY"),
    "OWEN S": person(True, "Owen, Stephen", "UKAPSOWE"),
    "Kenyon N": person(True, "Kenyon, Nathan", "UKAPNKEN"),
    "Rodgers K": person(True, "Rodgers, Kaleb", "UKAPKROD"),
    "Bruckshaw L": person(),
    "Minta N": person(),
}

converted_wb = Workbook()
converted_ws = converted_wb.create_sheet("Sheet1", 0)
converted_ws.append(
    [
        "Empl#",
        "Employee Name",
        "Day",
        "Time In",
        "Time Out",
        "Hours",
        "LL1",
        "LL2",
        "LL3",
        "LL4",
        "LL5",
        "LL6",
        "LL7",
        "Reg",
        "O/T",
    ]
)
row_num = 1
current_row = 0
employee_number = 0
employee = person()

for filepath in glob.glob(os.path.join(folder_path, "*.xls*")):
    wb = load_workbook(filename=filepath)
    print(filepath)
    sheet = wb['Sheet1']

    for row in sheet.iter_rows(values_only=True):
        if set(['Day', 'Daily Times', 'Basic', 'OT1', 'OT2', 'Hols', 'Sick', 'CC/AB', 'Daily Totals']) <= set(row):
            employee_number = sheet.cell(row=current_row, column=1).value
            employee = people[sheet.cell(row=current_row, column=2).value]
        current_row = current_row + 1
        try:
            start_time = datetime.strptime(f'{year}/{row[0]}','%Y/%a/%d/%b')
        except Exception as e:
            continue
        else:
            if not isinstance(row[1], time):
                continue
            if not isinstance(row[18], (int, float)):
                continue
            start_time = datetime.combine(start_time, row[1])
            end_time = start_time + timedelta(hours=row[18])
            if employee.laborcode == 'blank':
                print(f'Error finding employee: {employee_number}')
            converted_ws.append(
                [
                    employee_number,
                    employee.laborcode,
                    start_time.strftime('%A').upper(),
                    start_time.strftime('%Y-%m-%d %H:%M'),
                    end_time.strftime('%Y-%m-%d %H:%M'),
                    row[18] or 0,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    None,
                    row[10] or 0,
                    row[11] or 0,
                ]
            )

if current_row == 1:
    print('No .xls* files found, please check input folder')

converted_wb.save(save_path)
