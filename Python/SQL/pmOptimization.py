from datetime import datetime
import re

import pyodbc
from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError


conn = pyodbc.connect(
    "DRIVER={ODBC Driver 18 for SQL Server};SERVER=NSCANDACMAXSQL1\\MAXIMO;DATABASE=mxprod76;Trusted_Connection=yes;Encrypt=no;"
)

conn.setdecoding(pyodbc.SQL_CHAR, encoding="latin1")
conn.setdecoding(pyodbc.SQL_WCHAR, encoding="latin1")

save_path = f'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\SQL\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

cursor = conn.cursor()

cursor.execute(
    """
SELECT apm.eauditusername
    , apm.eaudittimestamp
    , apm.eaudittype
    , apm.siteid
    , apm.frequency
    , apm.frequnit
    , apm.pmnum
    , apm.pmuid
    , jp.JPDURATION
    , pm.status
FROM aws.maxprd.dbo.A_PM apm
left join aws.maxprd.dbo.pm on pm.pmuid = apm.pmuid
left join aws.maxprd.dbo.jobplan jp on pm.jpnum = jp.jpnum
WHERE eaudittype != 'D'
    AND apm.pmuid IN (
        SELECT apm.pmuid
        FROM aws.maxprd.dbo.a_pm
        WHERE eaudittype = 'u'
            AND datepart(year, eaudittimestamp) = 2024
        )
ORDER BY eaudittimestamp DESC
 """
)

wb = Workbook()
# get PM audit data
sheet = wb.create_sheet("pmData")
sheet.append(
    [
        "eauditusername",
        "eaudittimestamp",
        "eaudittype", #2
        "siteid",
        "frequency", #4
        "frequnit", #5
        "pmnum",
        "pmuid", #7
        "jpduration",
        "status"
    ]
)

CLEANR = re.compile("<.*?>")

pms = {}
updatedPms = {}
datePms = {}

for row in cursor:
    row = list(row)
    sheet.append(row)


updatedJobplans = {}
# get job plan audit data
cursor.execute(
"""
SELECT eauditusername
    , eaudittimestamp
    , eaudittype
    , siteid
    , jobplanid
    , jpnum
    , orgid
FROM aws.maxprd.dbo.a_jobplan
WHERE eaudittype = 'U'
    AND datepart(year, eaudittimestamp) = 2024
ORDER BY eaudittimestamp
 """
)
sheet = wb.create_sheet("jobplanData")
sheet.append(
    [
        "eauditusername",
        "eaudittimestamp",
        "eaudittype",
        "siteid",
        "jobplanid", #4
        "jpnum",
        "orgid",
    ]
)

CLEANR = re.compile("<.*?>")

for row in cursor:
    row = list(row)
    sheet.append(row)
    updatedJobplans[row[4]] = [row[3], row[5]]

# get jobtask audit data
cursor.execute(
    """
SELECT eauditusername
    , eaudittimestamp
    , eaudittype
    , siteid
    , orgid
    , jpnum
    , jobtaskid
FROM aws.maxprd.dbo.a_jobtask
WHERE eaudittype = 'U'
    AND datepart(year, eaudittimestamp) = 2024
ORDER BY eaudittimestamp
 """
)
sheet = wb.create_sheet("jobtasData")
sheet.append(
    [
        "eauditusername",
        "eaudittimestamp",
        "eaudittype",
        "siteid", #3
        "jobplanid",
        "jpnum", #5
        "orgid",
    ]
)

CLEANR = re.compile("<.*?>")

for row in cursor:
    row = list(row)
    sheet.append(row)
    updatedJobplans[row[4]] = [row[3], row[5]]

sheet = wb.create_sheet("jobplanUpdated")
for pm in updatedJobplans.values():
    sheet.append(list(pm))

conn.close()

wb.save(save_path)
