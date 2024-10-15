


from openpyxl import load_workbook

file_path = f'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\SQL\\04-09-2024_15-29-03.xlsx'

wb = load_workbook(filename=file_path)

sheet = wb['pmData']

frequnit = {
    'YEARS': 1,
    'WEEKS': 52,
    'MONTHS': 12,
    'DAYS': 365,
}

pms = {}
updatedPms = {}
datePms = {}

for row in sheet.iter_rows(values_only=True):
    if row[7] in pms.keys():
        if pms[row[7]] != [row[5], row[4]]: #if the frequency is now different
            try:
                newhours = datePms[row[7]][1] * (frequnit[pms[row[7]][0]] / pms[row[7]][1])
            except Exception:
                newhours = 'Error'
            try:
                oldhours = row[8] * (frequnit[row[5]] / row[4])
            except Exception:
                oldhours = 'Error'
            updatedPms[row[7]] = [row[3],row[6], datePms[row[7]][0].strftime('%Y-%m-%d'), oldhours, newhours, datePms[row[7]][2]]
        else: #if the freq is the same then shift the date back
            datePms[row[7]] = [row[1], row[8], row[9]] #date, duration
    else: # add pm to dictionary
        pms[row[7]] = [row[5], row[4]] #frequnit, freq
        datePms[row[7]] = [row[1], row[8], row[9]] #date, duration

sheet = wb.create_sheet("pmUpdated")
sheet.append(['siteid', 'pmnum', 'dateUpdated', 'oldHours', 'newHours', 'status'])
for key, item in updatedPms.items():
    # item.append(datePms[key].strftime('%Y-%m-%d'))
    sheet.append(item)

wb.save(file_path)