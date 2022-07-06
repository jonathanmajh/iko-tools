import os
import glob
from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

# ------ CHANGE ME ---------
folder_path = 'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\JDEItemRequestCombiner\\Files'
save_path = f'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\JDEItemRequestCombiner\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

batch_wb = Workbook()
alternativeDescription = batch_wb.create_sheet('Alt Description', 0)
alternativeDescription.append(['Item Number', 'Alt Desc 1', 'Alt Desc 2', 'Alt Search Text', 'Lang Code'])
itemBranch = batch_wb.create_sheet('ItemBranch', 0)
itemBranch.append(
    ['Item Number','Unit of Measure','G/L Class','Line Type','Commodity Class',
    'CatCode 8','CatCode 10','Maximo Item Type','Branch/Plant'])
itemMaster = batch_wb.create_sheet('ItemMaster', 0)
itemMaster.append(
    ['Item Number','Description1','Description2','Search Text','Unit of Measure',
    'G/L Class','Line Type','Commodity Class','CatCode 8','CatCode 10','Maximo Item Type'])

branch_row = []
master_row = []

for filepath in glob.glob(os.path.join(folder_path, '*.xls*')):
    wb = load_workbook(filename = filepath)
    print(filepath)
    for sheet in wb.worksheets:
        branch_row.append(sheet.cell(row=4, column=5).value)
        master_row.append(sheet.cell(row=4, column=5).value)
        master_row.append(sheet.cell(row=5, column=5).value)
        master_row.append(sheet.cell(row=6, column=5).value)
        master_row.append(sheet.cell(row=7, column=5).value)
        for x in range(8, sheet.max_row + 1):
            if sheet.cell(row=x, column=1).value == 'Maximo Description:':
                alternativeDescription.append(
                [sheet.cell(row=4, column=5).value, # itemnum
                    sheet.cell(row=x, column=5).value, #d esc1
                    sheet.cell(row=x+1, column=5).value, # desc2
                    sheet.cell(row=x+2, column=5).value, # desc3
                    sheet.cell(row=x-1, column=1).value, # lang code
                ]
            )
            elif sheet.cell(row=x, column=1).value == 'Issue Unit:':
                branch_row.append(sheet.cell(row=x, column=5).value)
                branch_row.append(sheet.cell(row=x+6, column=5).value)
                branch_row.append('B1')
                branch_row.append(sheet.cell(row=x+1, column=5).value)
                branch_row.append(sheet.cell(row=x+8, column=5).value)
                branch_row.append(sheet.cell(row=x+9, column=5).value)
                branch_row.append('I')
                branch_row.append(sheet.cell(row=x+7, column=5).value)

                master_row.append(sheet.cell(row=x, column=5).value)
                master_row.append(sheet.cell(row=x+6, column=5).value)
                master_row.append('N1')
                master_row.append(sheet.cell(row=x+1, column=5).value)
                master_row.append('ITM')
                master_row.append('ITM')
                master_row.append('I')
                itemMaster.append(master_row)
                itemBranch.append(branch_row)
                branch_row = []
                master_row = []
                break


batch_wb.save(save_path)