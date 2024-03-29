import os
import glob
from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

# ------ CHANGE ME ---------
folder_path = 'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\JDEItemRequestCombiner\\Files'
save_path = f'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\JDEItemRequestCombiner\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

batch_wb = Workbook()
alternativeDescription = batch_wb.create_sheet('Alt Description', 0)
alternativeDescription.append(['Item Number', 'Alt Desc 1', 'Alt Desc 2', 'Alt Search Text'])
itemBranch = batch_wb.create_sheet('ItemBranch', 0)
itemBranch.append(
    ['Item Number','Unit of Measure','G/L Class','Line Type','Commodity Class',
    'CatCode 8','CatCode 10','Maximo Item Type','Branch/Plant'])
itemMaster = batch_wb.create_sheet('ItemMaster', 0)
itemMaster.append(
    ['Item Number','Description1','Description2','Search Text','Unit of Measure',
    'G/L Class','Line Type','Commodity Class','CatCode 8','CatCode 10','Maximo Item Type'])

for filepath in glob.glob(os.path.join(folder_path, '*.xls*')):
    wb = load_workbook(filename = filepath)
    print(filepath)
    for sheet in wb.worksheets:
        itemBranch.append(
            [
                sheet.cell(row=4, column=5).value,
                sheet.cell(row=11, column=5).value,
                sheet.cell(row=17, column=5).value,
                'B1',
                sheet.cell(row=12, column=5).value,
                sheet.cell(row=20, column=5).value,
                sheet.cell(row=21, column=5).value,
                'I',
                sheet.cell(row=19, column=5).value,
            ]
        )
        itemMaster.append(
            [
                sheet.cell(row=4, column=5).value,
                sheet.cell(row=5, column=5).value,
                sheet.cell(row=6, column=5).value,
                sheet.cell(row=7, column=5).value,
                sheet.cell(row=11, column=5).value, #uom
                sheet.cell(row=17, column=5).value, #gl class
                'N1',
                sheet.cell(row=12, column=5).value, #commod class
                'ITM',
                'ITM',
                'I'
            ]
        )
        if (isinstance(sheet.cell(row=8, column=2).value, str) > 0):
            alternativeDescription.append(
                [
                    sheet.cell(row=4, column=5).value,
                    sheet.cell(row=8, column=5).value,
                    sheet.cell(row=9, column=5).value,
                    sheet.cell(row=10, column=5).value,
                ]
            )

batch_wb.save(save_path)