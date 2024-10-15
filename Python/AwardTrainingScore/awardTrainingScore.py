import os
import glob
import re
import csv
from datetime import datetime, timedelta, time

from openpyxl import Workbook
from openpyxl import load_workbook

from dataclasses import dataclass
# ------ CHANGE ME ---------
open_path = "C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\AwardTrainingScore\\Copy of Activity Completion Records Report - YTD October 2024.xlsx"
save_path = f'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\AwardTrainingScore\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.csv'
year = 2024

siteUsers = {}
# site: userid : []
users = {}
# userid: siteid
sites = {}
# siteid: []
completedDate = datetime

with open("C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\AwardTrainingScore\\PeopleWithSiteId.csv") as f:
    reader = csv.reader(f, delimiter=',')
    data = list(reader)

# load user mapping
for row in data:
    if (row[1] != 'Ignore'):
        siteUsers[row[1].upper()] = {}
    users[row[0].upper()] = row[1]

wb = load_workbook(filename=open_path)
sheet = wb['Activity Completion Records Rep']

for row in sheet.iter_rows():
    if row[0] is not None:
        if row[0].value in ('User', 'Credited'):
            row[5].value = row[5].value.upper()
            if row[5].value not in users.keys():
                print(f'Missing: {row[5].value}')
                continue
            if users[row[5].value] == 'Ignore':
                continue
            if row[5].value not in siteUsers[users[row[5].value]].keys():
                siteUsers[users[row[5].value]][row[5].value] = [0,0,0,0,0,0,0,0,0,0,0,0]
            try:
                if (type(row[6].value) == datetime):
                    completedDate = row[6].value
                else:
                    completedDate = datetime.strptime(row[6].value, '%m/%d/%Y')
            except Exception as e:
                print(f'Invalid Date: "{row[6].value}" on row: {row[6].row}')
            else:
                # siteUsers[siteid][userid][month] = minutesCompleted
                siteUsers[users[row[5].value]][row[5].value][completedDate.month - 1] = siteUsers[users[row[5].value]][row[5].value][completedDate.month - 1] + int(row[8].value)
                if siteUsers[users[row[5].value]][row[5].value][completedDate.month - 1] >= 60:
                    siteUsers[users[row[5].value]][row[5].value][completedDate.month - 1] = 60

for site in siteUsers.keys():
    sites[site] = [sum(x)/60 for x in zip(*siteUsers[site].values())]

with open(save_path, 'w', newline='') as f:
    writer = csv.writer(f)
    writer.writerow(['year', 'month', '', 'score', '', 'lookup', 'siteid'])
    for  site in sites.keys():
        for month, data in enumerate(sites[site]):
            writer.writerow([year, month + 1, '' ,data / len(siteUsers[site].keys()), '' , f'{site}{year}{month+1}', site])

print('done')