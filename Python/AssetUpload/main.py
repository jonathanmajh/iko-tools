import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl import Workbook

# ------ CHANGE ME ---------
folder_path = 'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\workOrderToJson\\'
save_path = f'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\workOrderToJson\\PiranaWOConversion.xlsm'

source_wb = load_workbook(filename = save_path, read_only=True, data_only=True)
ws = source_wb["ProcessedPiranaWOs"]

upload_data = '[\n'

for row in ws.iter_rows():
    upload_data = upload_data + '{\n'
    upload_data = f'{upload_data}"wonum": "{row[20].value}",\n'
    upload_data = f'{upload_data}"description": "{row[22].value}",\n'
    temp = row[19].value.replace("\n", "\\n")
    temp = temp.replace('"', '\\"')
    upload_data = f'{upload_data}"description_longdescription": "{temp}",\n'
    upload_data = f'{upload_data}"siteid": "KLU",\n'
    upload_data = f'{upload_data}"orgid": "IKO-EU",\n'
    upload_data = f'{upload_data}"woclass": "WORKORDER",\n'
    upload_data = f'{upload_data}"assetnum": "{row[4].value}",\n'
    if (row[4].value != None and row[4].value != '#N/A'):
        upload_data = f'{upload_data}"supervisor": "{row[8].value}",\n'
    if (row[13].value != None):
        upload_data = f'{upload_data}"iko_downtime": "{row[17].value}",\n'
        upload_data = f'{upload_data}"iko_desc1": ".",\n'
    if (row[14].value != None and row[23].value == "N"):
        upload_data = f'{upload_data}"jpnum": "{row[21].value}",\n'
    if (row[3].value != None and type(row[3].value) == datetime):
        upload_data = f'{upload_data}"reportdate": "{str(row[7].value).replace(" ", "T")}",\n'
    if (row[5].value != None and type(row[5].value) == datetime):
        upload_data = f'{upload_data}"schedstart": "{str(row[10].value).replace(" ", "T")}",\n'
    if (row[6].value != None and type(row[6].value) == datetime):
        upload_data = f'{upload_data}"actstart": "{str(row[15].value).replace(" ", "T")}",\n'
    if (row[7].value != None and type(row[7].value) == datetime):
        upload_data = f'{upload_data}"actfinish": "{str(row[16].value).replace(" ", "T")}",\n'
    upload_data = f'{upload_data}"status": "WPLAN"\n'
    upload_data = upload_data + '},\n'

upload_data = upload_data + ']'

with open('C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\workOrderToJson\\readme.json', 'w') as f:
    f.write(upload_data)