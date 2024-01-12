import requests
import csv

from openpyxl import load_workbook
from io import StringIO

wonums = {}

wb1 = load_workbook(
    filename="C:\\Users\\majona\\Documents\\Pirana\\WOsInMaximo.xlsx",
    read_only=True,
    data_only=True
)
sheet1 = wb1["S1"]
for row in sheet1.iter_rows(min_row=2, values_only=True):
    wonums[row[0]] = row[0]


baseUrl = "https://test.manage.test.iko.max-it-eam.com/maximo/api/os/xxx?action=importfile&lean=1"
apiKey = "smqtrblh2ofemdkgl0roh1vbod7k6ph3ua9kgg3h"
headers = {"apikey": apiKey, "Content-Type": "text/plain"}
url = baseUrl.replace("xxx", "IKO_WO")

wb = load_workbook(
    filename="C:\\Users\\majona\\Documents\\Pirana\\PiranaWOConversion.xlsm",
    read_only=True,
    data_only=True
)
sheet = wb["ClosedWOsPart2"]

head = [
    "WONUM",
    "DESCRIPTION",
    "REPORTDATE",
    "SUPERVISOR",
    "SCHEDSTART",
    "ACTSTART",
    "ACTFINISH",
    "ORGID",
    "SITEID",
    "WOCLASS",
    "STATUS",
    "ASSETNUM",
    "IKO_DOWNTIME",
    "JPNUM",
    "IKO_DESC1",
    "WOLONGDESC",
]

with open('result.txt', 'w', newline='') as f1:
    writer = csv.writer(f1)
    for row in sheet.iter_rows(min_row=620, values_only=True):
        if row[0] in wonums.keys():
            continue
        f = StringIO()
        csv.writer(f).writerow(head)
        row2 = []
        for cell in row:
            if cell is None:
                row2.append('')
            elif str(type(cell)) == "<class 'datetime.datetime'>":
                row2.append(cell.strftime('%Y-%m-%dT%H:%M:%S'))
            else:
                row2.append(str(cell))
        csv.writer(f).writerow(row2)
        data = f.getvalue()
        print(data)
        req = requests.post(url, headers=headers, data=data)
        writer.writerow([row[0],req.text])
        print([row[0], req.text])