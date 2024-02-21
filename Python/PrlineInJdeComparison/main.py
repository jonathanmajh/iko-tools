from datetime import datetime
import re
import easygui

import pyodbc
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError

sites = {
    200100: "AA",
    110200: "ANT",
    200140: "BA",
    191700: "BL",
    190800: "CA",
    140100: "CAM",
    160100: "COM",
    190600: "GC",
    200320: "GE",
    200120: "GH",
    200160: "GI",
    200200: "GJ",
    200300: "GK",
    200340: "GM",
    200220: "GP",
    200400: "GR",
    191000: "GS",
    191200: "GV",
    200700: "GX",
    120200: "KLU",
    180100: "PBM",
    140300: "RAM",
}

jdelines = {}

filename = easygui.fileopenbox(title="Select JDE Export File")
wb = load_workbook(filename=filename)
sheet = wb["MRO Orders"]

for row in sheet.iter_rows():
    if row[16].value not in sites:  # check if we are interested in this site
        continue
    siteid = sites[row[16].value]
    if siteid not in jdelines:  # check siteid
        jdelines[siteid] = {}
    if row[12].value not in jdelines[siteid]:  # check prnum
        jdelines[siteid][row[12].value] = {}
    jdelines[siteid][row[12].value][row[30].value] = row[3].value

conn = pyodbc.connect(
    "DRIVER={ODBC Driver 18 for SQL Server};SERVER=NSCANDACMAXSQL1\MAXIMO;DATABASE=mxprod76_intermed;Trusted_Connection=yes;Encrypt=no;"
)

conn.setdecoding(pyodbc.SQL_CHAR, encoding="latin1")
conn.setdecoding(pyodbc.SQL_WCHAR, encoding="latin1")

save_path = f'C:\\Users\\majona\\Documents\\Code\\iko-tools\\Python\\PrlineInJdeComparison\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

cursor = conn.cursor()

cursor.execute(
    """
SELECT pr.siteid
    , pr.prnum
    , pr.issuedate
    , pr.status
    , pl.itemnum
    , item.description
    , pl.orderqty
    , pl.ponum
FROM prline pl
LEFT JOIN pr
    ON pr.siteid = pl.siteid
        AND pl.prnum = pr.prnum
LEFT JOIN item
    ON item.itemnum = pl.itemnum
WHERE pr.STATUS = 'issued'
    AND pr.issuedate > '2023-01-01'
    AND pr.issuedate < '2024-01-01'
 """
)

wb = Workbook()
sheet = wb.create_sheet("Data")
sheet.append(
    [
        "siteid",
        "prnum",
        "issuedate",
        "status",
        "itemnum",
        "orderqty",
        "ponum",
        "JDEponum",
    ]
)

for row in cursor:
    row2 = list(row)
    if (row2[0] in jdelines): # check site
        if (f'0{row2[1]}' in jdelines[row2[0]]): # check prnum
            if (row2[4] in jdelines[row2[0]][f'0{row2[1]}']): # check item
                continue
    row2.append('Missing')
    sheet.append(row2)


conn.close()

wb.save(save_path)
