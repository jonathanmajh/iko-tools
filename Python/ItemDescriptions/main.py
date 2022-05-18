import re
import csv

from openpyxl import load_workbook
from openpyxl import Workbook

manus = {}
with open("Python\\ItemDescriptions\\companies.csv", encoding="utf8") as f:
    for line in f:
        manus[line.strip().split(",").pop(0)] = 1

items = {
    402: {},
    407: {},
    408: {},
    403: {},
    406: {},
    490: {},
    405: {},
    401: {},
    420: {},
    404: {},
    409: {},
    500: {},
    415: {},
    650: {},
    7660: {},
    430: {},
    "NULL": {},
    425: {},
    510: {},
    511: {},
    870: {},
    600: {},
    860: {},
    7405: {},
    700: {},
    750: {},
}

wb = load_workbook(filename="Python\\ItemDescriptions\\ItemDescriptions.xlsx")
sheet = wb['Sheet1']
for row in sheet.iter_rows(values_only=True):
    description = row[1].strip().split(",")
    item_type = description.pop(0)
    if not (item_type in items[row[2]]):
        items[row[2]][item_type] = {}
    for word in description:
        match = re.search("[0-9]+", word)
        if not match and not (word in manus):
            if not (word in items[row[2]][item_type]):
                items[row[2]][item_type][word] = 1
            else:
                items[row[2]][item_type][word] = items[row[2]][item_type][word] + 1

with open("Python\\ItemDescriptions\\result.csv", "w", encoding="UTF8", newline="") as f:
    write = csv.writer(f)
    for c_group in items:
        for item in items[c_group]:
            for word in items[c_group][item]:
                row = [c_group, item, word, items[c_group][item][word], f"{c_group}{item}{items[c_group][item][word]}"]
                write.writerow(row)
# write as {item_type}{count} < this should be sortable while keeping item_type grouped
# print(items)
