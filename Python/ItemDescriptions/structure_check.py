from openpyxl import load_workbook
from openpyxl import Workbook

import re
import csv
from dataclasses import dataclass

file_path = 'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\ItemDescriptions\\Feb28-2022ItemMaster.xlsx'
save_file = 'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\ItemDescriptions\\Feb28-2022ItemMaster_analysis.xlsx'

data = {}
description = ''
tree = ''
parent = ''

manus = {}
with open('companies.csv', mode='r', encoding='utf-8-sig') as f:
    for line in f:
        manus[line.split(',')[0].strip()] = 1

@dataclass
class descriptor:
    phrase: str
    parent: str
    count: int
    level: int

data_wb = load_workbook(filename=file_path)
ws = data_wb['Sheet1']

for row in ws.iter_rows(min_row=2):
    tree = ''
    description = row[1].value.strip().split(',')
    parent = ''
    level = 0
    for word in description:
        match = re.search('[0-9]+', word)
        if not match and not (word in manus):
            level = level + 1
            if len(tree) > 0:
                tree = tree + ',' + word
            else:
                tree = word
            if not(tree in data):
                data[tree] = descriptor(word, parent, 1, level)
            else:
                data[tree].count = data[tree].count + 1
            parent = tree

output = Workbook()
analysis = output.create_sheet('Analysis', 0)
analysis.append(['tree','descriptor', 'parent', 'count', 'level'])
for item in data:
    analysis.append([item, data[item].phrase, data[item].parent, data[item].count, data[item].level])

output.save(save_file)

print('complete')