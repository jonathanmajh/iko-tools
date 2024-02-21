import requests
import json

from openpyxl import load_workbook

headers = {
    "apikey": "smqtrblh2ofemdkgl0roh1vbod7k6ph3ua9kgg3h",
    "Patchtype": "MERGE",
    "X-Method-Override": "PATCH",
    "Content-Type": "application/json",
}

filepath = "C:\\Users\\majona\\Documents\\Pirana\\PiranaWOConversion.xlsm"
wb = load_workbook(filename=filepath)
sheet = wb["Sheet1"]

for row in sheet.iter_rows():
    if len(row[0].value) > 0:
      url = f'https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/MXAPIWODETAIL?lean=1&oslc.where=wonum="{row[1].value}" and siteid="KLU"&oslc.select=*'
      response = requests.get(url, headers=headers)
      wo_id = response.json()["member"][0]["href"].split("/")[-1]
      print(wo_id)

      url = f"https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/mxapiwodetail/{wo_id}/woworklog?oslc.select=*&oslc.pageSize=200&collectioncount=1&ignorecollectionref=1&relativeuri=1&interactive=1&lean=1&internalvalues=1"

      payload = json.dumps(
          [
              {
                  "createby": "MAJONA",
                  "createdate": "2024-01-23T00:00:00",
                  "logtype": "WORK",
                  "description": "Work Done Notes",
                  "description_longdescription": row[0].value,
              }
          ]
      )
      response = requests.request("POST", url, headers=headers, data=payload)
      print(response.text)
