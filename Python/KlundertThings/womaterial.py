import requests
import json

from openpyxl import load_workbook

headers = {
    "apikey": "smqtrblh2ofemdkgl0roh1vbod7k6ph3ua9kgg3h",
    "Patchtype": "MERGE",
    "X-Method-Override": "PATCH",
    "Content-Type": "application/json",
}

filepath = "C:\\Users\\majona\\Documents\\Pirana\\PiranaWOConversion.xlsm"
wb = load_workbook(filename=filepath)
sheet = wb["Sheet2"]

for row in sheet.iter_rows():
    url = f'https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/MXAPIWODETAIL?lean=1&oslc.where=wonum="{row[3].value}" and siteid="KLU"&oslc.select=*'
    response = requests.get(url, headers=headers)
    wo_id = response.json()["member"][0]["href"].split("/")[-1]
    print(wo_id)

    url = f"https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/mxapiwodetail/{wo_id}/uxshowactualmaterial?oslc.select=*&oslc.pageSize=2&collectioncount=1&ignorecollectionref=1&relativeuri=1&interactive=1&lean=1&internalvalues=1"

    payload = json.dumps(
        [
            {
                "itemnum": row[2].value,
                "storeloc": "KD1",
                "positivequantity": row[0].value,
                # "binnum": "20-A5-33",
                # "curbal": 5,
                "issuetype": "ISSUE",
                # "rotassetnum": "",
                # "location": "L-B3630",
                # "locdesc": "IKO Klundert Maintenance Storeroom",
                "description": row[1].value,
                # "taskid": "",
                "actualdate": "2024-01-23T00:00:00",
            }
        ]
    )
    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)
