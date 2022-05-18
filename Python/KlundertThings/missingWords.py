import os
import glob
from datetime import datetime, timedelta, time

from openpyxl import load_workbook
from openpyxl import Workbook

# ------ CHANGE ME ---------
folder_path = "C:\\Users\\majona\\GitHub\\iko-tools\\Python\\KlundertThings\\input"
save_path = f'C:\\Users\\majona\\GitHub\\iko-tools\\Python\\KlundertThings\\{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx'

for filepath in glob.glob(os.path.join(folder_path, "*.xls*")):
    wb = load_workbook(filename=filepath)
    print(filepath)
    sheet = wb['Sheet2']

    #  use set subtraaction to find left out words set([1,2,3,3]) - set([1,3])
    row_num = 2
    for row in sheet.iter_rows(values_only=True, min_row=2):
        if row[11] is not None:
            missing = []
            original = f'{row[3]} {row[4]}'
            new = row[11].upper()
            original = set(original.split(' '))
            for word in original:
                if word.upper() not in new:
                    missing.append(word)
            sheet.cell(row = row_num, column=7, value=','.join(missing))
        row_num = row_num + 1

wb.save(save_path)
