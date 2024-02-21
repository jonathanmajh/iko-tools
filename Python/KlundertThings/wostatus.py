import requests
import json

from openpyxl import load_workbook

headers = {
    "apikey": "smqtrblh2ofemdkgl0roh1vbod7k6ph3ua9kgg3h",
    "Patchtype": "MERGE",
    "X-Method-Override": "PATCH",
    "Content-Type": "application/json",
}

filepath = "C:\\Users\\majona\\Documents\\Pirana\\PiranaWOtoClose.xlsx"
wb = load_workbook(filename=filepath)
sheet = wb["Sheet1"]

for row in sheet.iter_rows():
    url = f'https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/MXAPIWODETAIL?lean=1&oslc.where=wonum="{row[0].value}" and siteid="KLU"&oslc.select=*'
    response = requests.get(url, headers=headers)
    wo_id = response.json()["member"][0]["href"].split("/")[-1]
    print(wo_id)

    url = f"https://prod.manage.prod.iko.max-it-eam.com/maximo/api/os/mxapiwodetail/{wo_id}?action=wsmethod%3AchangeStatus&interactive=false&lean=1&relativeuri=1&internalvalues=1"

    payload = json.dumps(
        {
            "date": "2024-01-26T00:00:00",
            "memo": "",
            "status": "CLOSE",
        }
    )
    response = requests.request("POST", url, headers=headers, data=payload)
    print(response.text)
