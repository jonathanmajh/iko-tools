import os
import glob

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import csv

asset_tasks = {}

for filepath in glob.glob(os.path.join('C:\\Users\\majona\\Desktop\\AA PMs', '*.xls*')):
    wb = load_workbook(filename = filepath)
    print(filepath)
    ws = wb["Main"]

    new_row = False
    # header_row = 0
    assetnum = None

    for row in ws.iter_rows():
        if row[0].value == 'PM Asset/ Parent (Route)*:':
            new_row = True
            header_row = row[0].row
        elif new_row:
            new_row = False
            route = row[1].value
            siteid = row[3].value
            assetnum = row[0].value
        elif assetnum:
            if row[0].row == header_row + 2:
                pm_num = row[8].value
            elif row[6].data_type == 'n' and row[7].value is not None:
                description = row[7].value.replace("(", " ").replace(")", " ").replace("  ", " ")
                if row[7].value != 'General' and row[7].value != 'Completion' and route == 'None' and row[4].value is None:
                    if asset_tasks.get(assetnum) is None:
                        asset_tasks[assetnum] = [[pm_num, description.split(),filepath]]
                    else:
                        asset_tasks[assetnum].append([pm_num, description.split(),filepath])
                elif route == 'Child' and row[4].value is not None:
                    assetnum = row[4].value
                    if asset_tasks.get(assetnum) is None:
                        asset_tasks[assetnum] = [[pm_num, description.split(),filepath]]
                    else:
                        asset_tasks[assetnum].append([pm_num, description.split(),filepath])

f = open('C:\\Users\\majona\\Documents\\GitHub\\iko-tools\\Python\\Assets_JobTasks\\asset_tasks.csv', 'w')

writer = csv.writer(f)

for asset in asset_tasks:
    for pm in asset_tasks[asset]:
        temp = [pm[2],pm[0],asset]
        for word in pm[1]:
            temp.append(word)
        # pm[0].append(pm[1])
        writer.writerow(temp)

f.close()
