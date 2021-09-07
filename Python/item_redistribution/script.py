from openpyxl import load_workbook, Workbook

wb = load_workbook(filename='C:\\Users\\majona\\Documents\\GitHub\\iko-tools\\Python\\item_redistribution\\wilmington_parts.xlsx', read_only=True)

sheet_transact = wb['site item with transaction']

sheet_balance = wb['site item with balance']

sheet_inventory = wb['item in site inventory']

sheet_vendors = wb['all vendors']

sheet_sites = wb['Sites']

ga_item_details = wb['vendor + prices at GA']

ga_item_balance = wb['GA items with balance']

# sheet_result = wb.create_sheet(title='results')

order = ['BL','GV','GS','CA','GC']

ca_order = ['AA','GH','GI','GK','GR','GX','BA','GE','GM']

dict_transact = {}

dict_balance = {}

dict_inventory = {}

dict_vendors = {}

dict_sites = {}

dict_item_details = {}

for row in sheet_vendors.iter_rows(min_row=1, max_col=3, max_row=sheet_vendors.max_row, values_only=True):
    dict_vendors[str(row[0])]=str(row[2])

for row in sheet_sites.iter_rows(min_row=1, max_col=3, max_row=sheet_sites.max_row, values_only=True):
    dict_sites[str(row[0])]=str(row[1])

print('done companies')

for row in sheet_transact.iter_rows(min_row=1, max_col=6, max_row=sheet_transact.max_row, values_only=True):
    if str(row[0]) in dict_transact:
        dict_transact[str(row[0])].append(str(row[1]))
    else:
        dict_transact[str(row[0])] = [str(row[1])]

print('done transacts')

for row in sheet_balance.iter_rows(min_row=1, max_col=3, max_row=sheet_balance.max_row, values_only=True):
    if str(row[0]) in dict_balance:
        dict_balance[str(row[0])].append(str(row[2]))
    else:
        dict_balance[str(row[0])] = [str(row[2])]

print('done balance')

for row in sheet_inventory.iter_rows(min_row=1, max_col=2, max_row=sheet_inventory.max_row, values_only=True):
    if str(row[0]) in dict_inventory:
        dict_inventory[str(row[0])].append(str(row[1]))
    else:
        dict_inventory[str(row[0])] = [str(row[1])]

print('done inventory')

for row in ga_item_details.iter_rows(min_row=1, max_col=5, max_row=ga_item_details.max_row, values_only=True):
    dict_item_details[str(row[0])] = [str(row[1]), str(row[3])]

print('done item details')

all_inv_dicts = [dict_transact, dict_balance, dict_inventory]

wb = Workbook()
ws1 = wb.active
ws1.title = 'Results'
ws1.append([])
ws1.append([])
ws1.append(['Item Number', 'Item Name', 'Quantity', 'UOM', 'Vendor ID', 'Unit Price', 'Vendor Name', None,'BL','GV','GS','CA','GC','BL','GV','GS','CA','GC','BL','GV','GS','CA','GC', None,'AA','GR','GX','GH','GI','GJ','GK','GM','BA','GE', 'AA','GR','GX','GH','GI','GJ','GK','GM','BA','GE','AA','GR','GX','GH','GI','GJ','GK','GM','BA','GE'])

for row in ga_item_balance.iter_rows(min_row=2, max_col=4, max_row=ga_item_balance.max_row, values_only=True):
    result_row = []
    row = [str(row[0]), row[1], row[2], row[3]]
    result_row.extend(row)
    vendor = dict_item_details.get(row[0], False)
    if vendor: # get vendor info
        result_row.extend([vendor[0], vendor[1]])
        vendor = dict_vendors.get(vendor[0],None)
        result_row.append(vendor)
    else:
        result_row.extend([None,None,None])
    
    result_row.append(None)
    for inventory in all_inv_dicts: # first check US sites
        item_in_inventory = inventory.get(row[0], [])
        for site in order:
            if site in item_in_inventory:
                result_row.append(site)
            else:
                result_row.append(None)

    result_row.append(None)
    for inventory in all_inv_dicts: # now do CAD
        item_in_inventory = inventory.get(row[0], [])
        for site in ca_order:
            if site in item_in_inventory:
                result_row.append(site)
            else:
                result_row.append(None)

    ws1.append(result_row)

wb.save(filename='C:\\Users\\majona\\Documents\\GitHub\\iko-tools\\Python\\item_redistribution\\results.xlsx')

print('done')