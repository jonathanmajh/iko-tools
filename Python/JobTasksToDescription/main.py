import csv
from os import read

from openpyxl import load_workbook

jps = {}
tasks = {}
wb = load_workbook(filename = 'BL_Route_JobTasks2.xlsx')
ws = wb["BL_Route_JobTasks2"]

for row in ws.iter_rows():
    if (not(row[0].value in jps)):
        jps[row[0].value] = f'{row[4].value} - {row[1].value} - {row[5].value}'
        tasks[row[0].value] = [row[4].value]
    else:
        jps[row[0].value] = f'{jps[row[0].value]}\n{row[4].value} - {row[1].value} - {row[5].value}'
        if not (row[4].value in tasks[row[0].value]):
            tasks[row[0].value].append(row[4].value)

with open('result3.csv', 'w', encoding='UTF8', newline='') as f:
    write = csv.writer(f)
    for item in jps:
        if len(tasks[item]) > 1:
            row = [item, jps[item]]
            write.writerow(row)
        else:
            row = [item, tasks[item][0]]
            write.writerow(row)